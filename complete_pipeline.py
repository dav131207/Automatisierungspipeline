"""
complete_pipeline.py  –  ESF+ Vollautomatisierung
==================================================
Workflow:
  1. Excel Spalte E scannen → "Diözese: Unterlagen in Wolke hochgeladen"
  2. Vor-/Nachname aus Spalten G/H lesen
  3. Sdb- und Zg-PDFs finden (case-insensitiv)
  4. Ordner "Vorname Nachname" anlegen, PDFs hineinverschieben
  5. OCR: alle Felder aus Stammdatenblatt extrahieren
  6. Datenqualität prüfen → Spalte U (Ja/Nein) + Spalte V (Beschreibung)
  7. IDEA öffnen, neuen Teilnehmer anlegen
  8. IDEA-ID aus #tabs_enter... auslesen → Spalte R und S
  9. Spalte T (Interne Kennung) in IDEA eintragen
 10. Spalte E → "CÖ: IDEA Angelegt"

Abhängigkeiten:
    pip3 install pymupdf pytesseract Pillow openpyxl playwright
    playwright install chromium
"""

import sys, re, io, logging, configparser, time, shutil
from pathlib import Path

try:
    import fitz
    import pytesseract
    from pytesseract import Output
    from PIL import Image, ImageEnhance
except ImportError:
    sys.exit("FEHLER: pip3 install pymupdf pytesseract Pillow")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    sys.exit("FEHLER: pip3 install openpyxl")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    sys.exit("FEHLER: pip3 install playwright && playwright install chromium")

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)s]  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("pipeline.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

CONFIG_FILE = Path(__file__).parent / "config.ini"
SCREENSHOTS = Path(__file__).parent / "screenshots"

# Excel-Spalten (1-basiert)
COL_STATUS   = 5   # E
COL_NACHNAME = 7   # G
COL_VORNAME  = 8   # H
COL_ID_1     = 18  # R  – erster Teil der IDEA-ID  (z.B. 2624)
COL_ID_2     = 19  # S  – zweiter Teil              (z.B. 69611)
COL_INTERN   = 20  # T  – Interne Kennung (Formel aus R+S, wird von Excel berechnet)
COL_OK       = 21  # U  – Datenqualität: "Ja" / "Nein"
COL_FEHLER   = 22  # V  – Fehlerbeschreibung

STATUS_BEREIT   = "Diözese: Unterlagen in Wolke hochgeladen"
STATUS_ERLEDIGT = "CÖ: IDEA Angelegt"

# ═════════════════════════════════════════════════════════════════════════════
# OCR
# ═════════════════════════════════════════════════════════════════════════════

LEER = re.compile(
    r'^\s*(?:LJ|LI|LO|CL|CI|C1|CJ|Cl|sO|LY|U\b|C\b|O\b|0\b|\[\s*\]|L\s*J)',
    re.IGNORECASE
)
LEER_CB = re.compile(
    r'^(?:LJ|LI|LO|CL|CI|CJ|Oo|Cl|sO|UJ|LY|U\b|O\b|0\b)',
    re.IGNORECASE
)

def _img(seite, dpi=400, clip=None, k=2.0, up=1):
    pix = seite.get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72),
                           clip=clip, colorspace=fitz.csGRAY)
    img = ImageEnhance.Contrast(Image.open(io.BytesIO(pix.tobytes("png")))).enhance(k)
    if up > 1:
        img = img.resize((img.width*up, img.height*up), Image.LANCZOS)
    return img

def _o(seite, psm=4, k=1.8):
    return pytesseract.image_to_string(_img(seite, k=k), config=f"--psm {psm} --oem 3")

def _or(seite, x0, y0, x1, y1, up=2, psm=6):
    return pytesseract.image_to_string(
        _img(seite, clip=fitz.Rect(x0,y0,x1,y1), up=up),
        config=f"--psm {psm} --oem 3"
    )

def _b(t):
    """Bereinigt OCR-Artefakte am Anfang/Ende (inkl. Unterstriche, Linien)."""
    t = re.sub(r'^[\s|:\-!_]+|[\s|:\-!_:]+$', '', t).strip()
    # Entferne auch mehrfache Unterstriche (Formularlinien wie _____)
    t = re.sub(r'_+\s*$', '', t).strip()
    return t

def _cb(text, option):
    m = re.search(re.escape(option[:10]), text, re.I)
    if not m: return False
    return not LEER.match(text[m.end():m.end()+14].lstrip(" |:"))

# Erweiterte Checkmark-Muster (verschiedene OCR-Varianten von ☒):
# Xl, IX, [X, &l, x], X], Xl usw.
_CHECK_PAT = re.compile(
    r'[☒⊠Xx]|\[X|IX\]|X\]|Xl\b|&l\b|\[Xl',
    re.IGNORECASE
)
_LEER_PAT = re.compile(
    r'^(?:LJ|LI|LO|CL|CI|C1|CJ|Cl|sO|UJ|LY|U\b|O\b|0\b|L\]|\[\s*\])',
    re.IGNORECASE
)

def _drei_box(text: str) -> str:
    """
    Erkennt Ja/Nein/Keine Angabe durch Ankerpunkt direkt nach dem Label.
    Gleiche Logik wie _box() bei Staatszugehörigkeit – robuster als _drei().
    """
    CHK = r'(?:' + _CHECK_PAT.pattern + r')'

    def box(pattern):
        m = re.search(pattern, text, re.I)
        if not m:
            return None
        return text[m.end():m.end()+6].lstrip(' |:\n\t')

    ja_box = box(r'Ja\s*[:\|U]\s*')
    ne_box = box(r'Nein\s*[:\|]?\s*')
    ka_box = box(r'Keine\s+Angabe\s*[:\|]?\s*')

    # Schritt 1: Positiver Checkmark direkt nach Label
    for wert, b in [('Keine Angabe', ka_box), ('Ja', ja_box), ('Nein', ne_box)]:
        if b and re.search(CHK, b, re.I) and not _LEER_PAT.match(b):
            return wert

    # Schritt 2: Ausschluss – welche Option hat KEINEN Leer-Marker?
    hat_leer = {
        'Ja':           ja_box is None or bool(_LEER_PAT.match(ja_box)),
        'Nein':         ne_box is None or bool(_LEER_PAT.match(ne_box)),
        'Keine Angabe': ka_box is None or bool(_LEER_PAT.match(ka_box)),
    }
    ohne = [w for w, l in hat_leer.items() if not l]
    if len(ohne) == 1:
        return ohne[0]

    return 'Keine Angabe'



def _zeile_ocr(page, stichwort: str, dpi: int = 500, toleranz: int = 22) -> str:
    """
    Findet mit PSM 6 (Bounding-Boxes) die Checkbox-Zeile für ein Stichwort.
    Bevorzugt Zeilen die auch "Ja" enthalten (echte Frage-Zeilen).
    """
    pix = page.get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72), colorspace=fitz.csGRAY)
    img = ImageEnhance.Contrast(
        Image.open(io.BytesIO(pix.tobytes('png')))).enhance(2.0)
    daten = pytesseract.image_to_data(
        img, config='--psm 6 --oem 3', output_type=Output.DICT)

    woerter = [
        {'text': daten['text'][i],
         'left': daten['left'][i],
         'mid_y': daten['top'][i] + daten['height'][i] // 2}
        for i in range(len(daten['text']))
        if daten['text'][i].strip()
    ]

    such = stichwort[:8].lower()
    kandidaten = []  # Alle y-Positionen wo Stichwort vorkommt
    for w in woerter:
        if such in w['text'].lower():
            kandidaten.append(w['mid_y'])

    if not kandidaten:
        return ''

    def zeile_bei(label_y):
        zeile = sorted(
            [w for w in woerter if abs(w['mid_y'] - label_y) <= toleranz],
            key=lambda w: w['left']
        )
        return ' '.join(w['text'] for w in zeile)

    # Bevorzuge Zeile die "Ja" enthält (Frage mit Checkbox)
    for y in kandidaten:
        z = zeile_bei(y)
        if re.search(r'\bJa\b', z, re.I):
            return z

    # Fallback: erste gefundene Zeile
    return zeile_bei(kandidaten[0])


def _check(text: str, label: str) -> str:
    """
    Erkennt ob Ja/Nein/Keine Angabe angekreuzt ist.
    Robuster als _drei(): erkennt auch [Xl, IX], &l usw.
    """
    # Suche Label + was direkt danach steht
    for opt, wert in [("Keine Angabe", "Keine Angabe"), ("Nein", "Nein"), ("Ja", "Ja")]:
        m = re.search(re.escape(opt), text, re.I)
        if not m: continue
        danach = text[m.end(): m.end()+10].lstrip(" |:\t")
        if _CHECK_PAT.match(danach) and not _LEER_PAT.match(danach):
            return wert
    return "Keine Angabe"  # Default

def _drei(text):
    jl = bool(re.search(r'Ja\s*[:\|U]\s*(?:[UC]?[JLO0]\b|\[\s*\])', text, re.I))
    nl = bool(re.search(r'Nein\s*[:\|]?\s*(?:[UC]?[JLO0]\b|\[\s*\])', text, re.I))
    jv = bool(re.search(r'Ja\s*[:\|U]\s*' + _CHECK_PAT.pattern, text, re.I))
    nv = bool(re.search(r'Nein\s*[:\|]?\s*' + _CHECK_PAT.pattern, text, re.I))
    kav = bool(re.search(r'Keine\s+Angabe\s*[:\|]?\s*' + _CHECK_PAT.pattern, text, re.I))
    if kav: return "Keine Angabe"
    if jv: return "Ja"
    if nv: return "Nein"
    if jl and nl: return "Keine Angabe"
    if jl: return "Nein"
    if nl: return "Ja"
    return "Keine Angabe"


# ─────────────────────────────────────────────────────────────────────────────
# Pixel-basierte Checkbox-Erkennung (offline, kein ML)
# Misst die Dichte dunkler Pixel rechts neben jedem Label.
# Angekreuzt ≈ 0.18–0.30  |  leer ≈ 0.02–0.15
# ─────────────────────────────────────────────────────────────────────────────
import numpy as np

_PIXEL_CACHE = {}

def _seiten_bild_und_woerter(page, dpi=300):
    """Seite als Graustufenbild + alle Wörter mit Pixel-Koordinaten (gecached)."""
    # Cache-Key über Dateinamen (id() kollidiert nach Garbage Collection!)
    key = (page.parent.name, page.number, dpi)
    if key in _PIXEL_CACHE:
        return _PIXEL_CACHE[key]
    pix = page.get_pixmap(matrix=fitz.Matrix(dpi/72, dpi/72), colorspace=fitz.csGRAY)
    img = Image.open(io.BytesIO(pix.tobytes('png')))
    daten = pytesseract.image_to_data(img, config='--psm 6 --oem 3',
                                      output_type=Output.DICT)
    woerter = []
    for i in range(len(daten['text'])):
        t = daten['text'][i].strip()
        if t:
            woerter.append({
                'text': t,
                'left': daten['left'][i], 'top': daten['top'][i],
                'width': daten['width'][i], 'height': daten['height'][i],
                'mid_y': daten['top'][i] + daten['height'][i] // 2,
            })
    _PIXEL_CACHE[key] = (img, woerter)
    return img, woerter


def _checkbox_dichte(img, x, y, breite=55, hoehe=45):
    """Anteil dunkler Pixel (Grauwert < 128) in der Checkbox-Region."""
    region = img.crop((max(0, x), max(0, y - hoehe // 2),
                       x + breite, y + hoehe // 2))
    arr = np.array(region)
    if arr.size == 0:
        return 0.0
    return float((arr < 128).sum()) / arr.size


def checkbox_optionen_pixel(page, stichworte, optionen=None, dpi=300,
                            zeilen_toleranz=35):
    """
    Findet die Zeile eines Stichworts, lokalisiert alle Options-Label
    (Ja/Nein/Keine Angabe) und misst die Pixel-Dichte rechts daneben.
    Gibt {Option: Dichte} zurück oder {} wenn nichts gefunden.
    """
    if optionen is None:
        optionen = {'ja': 'Ja', 'nein': 'Nein', 'angabe': 'Keine Angabe'}
    img, woerter = _seiten_bild_und_woerter(page, dpi)

    label_y = None
    for sw in (stichworte if isinstance(stichworte, list) else [stichworte]):
        for w in woerter:
            if sw.lower() in w['text'].lower():
                label_y = w['mid_y']
                break
        if label_y is not None:
            break
    if label_y is None:
        return {}

    zeile = [w for w in woerter if abs(w['mid_y'] - label_y) <= zeilen_toleranz]
    dichten = {}
    for w in zeile:
        t = w['text'].rstrip(':.').lower()
        if t in optionen:
            dichten[optionen[t]] = _checkbox_dichte(
                img, w['left'] + w['width'] + 5, w['mid_y'])
    return dichten


def drei_optionen_pixel(page, stichworte, schwelle=0.16):
    """
    Ja/Nein/Keine Angabe via Pixel-Dichte.
    Gewinner = höchste Dichte ÜBER der Schwelle. Sonst 'Keine Angabe'.
    """
    dichten = checkbox_optionen_pixel(page, stichworte)
    if not dichten:
        return None  # Label nicht gefunden → Aufrufer nutzt Fallback
    gewinner = max(dichten, key=dichten.get)
    if dichten[gewinner] >= schwelle:
        return gewinner
    return 'Keine Angabe'


def zwei_optionen_pixel(page, stichworte, schwelle=0.16):
    """Ja/Nein via Pixel-Dichte. None wenn Label nicht gefunden."""
    dichten = checkbox_optionen_pixel(
        page, stichworte, optionen={'ja': 'Ja', 'nein': 'Nein'})
    if not dichten:
        return None
    gewinner = max(dichten, key=dichten.get)
    if dichten[gewinner] >= schwelle:
        return gewinner
    # Beide unter Schwelle → vermutlich keines angekreuzt → Nein als Default
    return 'Nein'



def geburtsland_pixel(page, dpi=300, schwelle=0.16):
    """
    Geburtsland via Pixel-Dichte.
    Unterscheidet "Österreich:" (Option 1) von "Österreichs:" (Ende von
    "Außerhalb Österreichs:", Option 2) anhand des Wort-Endes.
    Gibt None zurück wenn die Zeile nicht gefunden wird.
    """
    img, woerter = _seiten_bild_und_woerter(page, dpi)
    gl_y = next((w['mid_y'] for w in woerter if 'eburtsland' in w['text']), None)
    if gl_y is None:
        return None
    zeile = [w for w in woerter if abs(w['mid_y'] - gl_y) <= 35]

    d_oe = d_aus = None
    for w in zeile:
        t = w['text'].rstrip(':.').lower()
        if re.match(r'^[o0öó6]sterreich$', t):
            d_oe = _checkbox_dichte(img, w['left'] + w['width'] + 5, w['mid_y'])
        elif re.match(r'^[o0öó6]sterreichs$', t):
            d_aus = _checkbox_dichte(img, w['left'] + w['width'] + 5, w['mid_y'])

    if d_oe is None and d_aus is None:
        return None
    if (d_oe or 0) >= (d_aus or 0) and (d_oe or 0) >= schwelle:
        return 'Österreich'
    if (d_aus or 0) >= schwelle:
        return 'Außerhalb Österreichs'
    return None  # beide unter Schwelle → Text-Fallback entscheidet



def elternteile_pixel(page, dpi=300, schwelle=0.16):
    """
    "Beide Elternteile im Ausland geboren" via Pixel-Dichte.
    Stufe 1: Wort-Anker 'Elternteile'/'geboren' (PSM 6 Wortliste).
    Stufe 2: Region zwischen 'Migrationshintergrund' und 'Staatszugehörigkeit'
             croppen, mit PSM 11 (sparse) Ja/Nein lokalisieren.
    Gibt 'Ja', 'Nein' oder None (Aufrufer nutzt Text-Fallback) zurück.
    """
    img, woerter = _seiten_bild_und_woerter(page, dpi)

    # ── Stufe 1: direkter Wort-Anker ────────────────────────────────────
    label_y = None
    for w in woerter:
        t = w['text'].lower()
        if 'elternteile' in t or t.startswith('geboren'):
            # Fußnote ausschließen (steht im unteren Seitendrittel)
            if w['mid_y'] < img.height * 0.6:
                label_y = w['mid_y']
                break
    if label_y is not None:
        zeile = [w for w in woerter if abs(w['mid_y'] - label_y) <= 35]
        ja_d = ne_d = None
        for w in zeile:
            t = w['text'].rstrip(':.').lower()
            if t == 'ja':
                ja_d = _checkbox_dichte(img, w['left'] + w['width'] + 5, w['mid_y'])
            elif t == 'nein':
                ne_d = _checkbox_dichte(img, w['left'] + w['width'] + 5, w['mid_y'])
        if ja_d is not None or ne_d is not None:
            if (ja_d or 0) >= (ne_d or 0) and (ja_d or 0) >= schwelle:
                return 'Ja'
            if (ne_d or 0) >= schwelle:
                return 'Nein'

    # ── Stufe 2: Region-Crop + PSM 11 ───────────────────────────────────
    mig_y   = next((w['mid_y'] for w in woerter
                    if 'igrationshint' in w['text'].lower()), None)
    staat_y = next((w['mid_y'] for w in woerter
                    if 'taatszug' in w['text'].lower()), None)
    if mig_y is None or staat_y is None or staat_y <= mig_y:
        return None

    region = img.crop((0, mig_y + 15, img.width, staat_y - 15))
    daten = pytesseract.image_to_data(region, config='--psm 11 --oem 3',
                                      output_type=Output.DICT)
    ja_d = ne_d = None
    for i in range(len(daten['text'])):
        t = daten['text'][i].strip().rstrip(':.').lower()
        if t in ('ja', 'nein'):
            x = daten['left'][i] + daten['width'][i] + 5
            y = daten['top'][i] + daten['height'][i] // 2
            d = _checkbox_dichte(region, x, y)
            if t == 'ja':
                ja_d = max(ja_d or 0, d)
            else:
                ne_d = max(ne_d or 0, d)

    if ja_d is None and ne_d is None:
        return None
    if (ja_d or 0) >= (ne_d or 0) and (ja_d or 0) >= schwelle:
        return 'Ja'
    if (ne_d or 0) >= schwelle:
        return 'Nein'
    return None


def parse_stammdatenblatt(pdf_pfad: Path) -> dict:
    """Vollständige OCR-Extraktion aus dem Stammdatenblatt."""
    log.info("    OCR: %s", pdf_pfad.name)
    doc = fitz.open(str(pdf_pfad))
    d   = {}

    # Seite 1: Vorname / Nachname mit Noise-Cleaning
    def _name(label, y0, y1):
        roh = _or(doc[0], 60, y0, 595, y1, up=3)
        m = re.search(label + r'\s*[:\|]\s*(.+)', roh, re.I)
        if not m:
            return ""
        n = m.group(1).strip()
        n = re.sub(r'[^\wäöüÄÖÜß\s\-\']', ' ', n, flags=re.UNICODE)
        n = re.sub(r'\s+', ' ', n).strip()
        return re.sub(r'^[\s\-\.\|]+|[\s\-\.\|]+$', '', n).strip()

    d["Vorname"]  = _name(r'Vorname',  536, 562)
    d["Nachname"] = _name(r'Nachname', 548, 580)

    # Geschlecht: Zeilen-OCR + Anchor (links→rechts: männlich→weiblich→nicht binär)
    gs_zeile = _zeile_ocr(doc[0], "Geschlecht")
    if not gs_zeile:
        gs_zeile = _or(doc[0], 60, 556, 595, 620)
    CHK_G = r"(?:" + _CHECK_PAT.pattern + r")"
    def _g_box(label):
        m = re.search(label, gs_zeile, re.I)
        if not m:
            return None
        return gs_zeile[m.end():m.end()+6].lstrip(" |:\n\t")
    mann_box = _g_box(r"(?:m.{0,2}nnlich|mannlich)[:\s]*")
    weib_box = _g_box(r"weiblich[:\s]*")
    nbp_box  = _g_box(r"nicht\s+bin[^\s]{0,3}[:\s]*")
    # Prüfung in Reihenfolge links→rechts (verhindert False Positives)
    if mann_box and re.search(CHK_G, mann_box, re.I) and not _LEER_PAT.match(mann_box):
        d["Geschlecht"] = "männlich"
    elif weib_box and re.search(CHK_G, weib_box, re.I) and not _LEER_PAT.match(weib_box):
        d["Geschlecht"] = "weiblich"
    elif nbp_box and re.search(CHK_G, nbp_box, re.I) and not _LEER_PAT.match(nbp_box):
        d["Geschlecht"] = "nicht binäre Person"
    else:
        d["Geschlecht"] = "männlich"  # Default

    # Seite 2: Kontakt / Adresse / Ausbildung / Erwerb
    s2 = _o(doc[1], psm=4)

    # SV-Nummer: auf allen Seiten suchen (Saleh hat SV auf Seite 1, Mustermann auf Seite 2)
    sv_keine = False
    sv_zahl  = ""
    for s_idx in range(min(3, len(doc))):
        s_txt = _o(doc[s_idx], psm=4) if s_idx != 1 else s2
        sv_bereich = re.search(r'Sozialversicherung.{0,200}', s_txt, re.DOTALL | re.I)
        if sv_bereich:
            sv_block = sv_bereich.group(0)
            if re.search(r'Keine\s+Angabe', sv_block, re.I):
                sv_keine = True
            m_sv = re.search(r'(\d{10})', sv_block)
            if m_sv:
                sv_zahl = m_sv.group(1)
            break
        # Fallback: allgemeine Suche auf der Seite
        if re.search(r'Keine\s+Angabe', s_txt[:400], re.I):
            sv_keine = True
            break
    d["SV-Nummer"]        = "Keine Angabe" if sv_keine else sv_zahl
    d["SV_Keine_Angabe"]  = sv_keine
    d["SV_Wert_Gefunden"] = bool(sv_zahl)

    for feld, muster in [
        ("Straße",  r'Stra[ßs]e\s*[:\|][ \t]*([^\n]{2,50})'),
        ("Hausnr",  r'Hausnr[./][^\n]*[:\|][ \t]*([^\n]{2,20})'),
        ("PLZ",     r'PLZ\s*[:\|][ \t]*(\d{4,5})'),
        ("Ort",     r'Ort\s*[:\|][ \t]*([^\n]{2,60})'),
        ("Land",    r'Land\s*[:\|][ \t]*([^\n]{2,60})'),
    ]:
        m = re.search(muster, s2)
        d[feld] = m.group(1).strip() if m else ""

    # Telefon: Versuch 1 – PSM 4 Volltext (funktioniert bei Mustermann-Form)
    m = re.search(r'Telefonnummer\s*[:\|]\s*([\d/+\s\-]{6,30})', s2)
    if m:
        d["Telefon"] = m.group(1).strip()
    else:
        # Versuch 2 – PSM 7 (Einzel-Zeilen) bei 600 DPI, breiter Y-Bereich
        # Deckt verschiedene Formularversionen ab (y=155-290)
        tel_gefunden = ""
        for y1, y2 in [(155, 180), (200, 230), (230, 260), (235, 265), (250, 280)]:
            try:
                clip = fitz.Rect(60, y1, 595, y2)
                pix  = doc[1].get_pixmap(
                    matrix=fitz.Matrix(600/72, 600/72), clip=clip,
                    colorspace=fitz.csGRAY
                )
                img  = ImageEnhance.Contrast(
                    Image.open(io.BytesIO(pix.tobytes("png")))
                ).enhance(2.0)
                img  = img.resize((img.width*2, img.height*2), Image.LANCZOS)
                zeile = pytesseract.image_to_string(
                    img, config="--psm 7 --oem 3"
                ).strip()
                mt = re.search(r'(0\d[\d/\s\-]{5,15})', zeile)
                if mt and len(mt.group(1).replace(" ","").replace("-","")) >= 7:
                    tel_gefunden = mt.group(1).strip().split()[0]  # erstes Token
                    break
            except Exception:
                continue
        d["Telefon"] = tel_gefunden

    m = re.search(r'([a-zA-Z0-9_.+\-]+@[a-zA-Z0-9\-]+\.[a-zA-Z]{2,})', s2)
    d["Email"] = m.group(1) if m else ""

    alle = re.findall(r'\d{2}[./]\d{2}[./]\d{4}', s2)
    d["Geburtsdatum"] = alle[0] if alle else ""

    # Geburtsland: Pixel-Dichte primär, Text-Anker Fallback
    gl_pixel = geburtsland_pixel(doc[1])
    if gl_pixel is not None:
        d["Geburtsland"] = gl_pixel
    else:
        gl_m = re.search(r'Geburtsland[^\n]*', s2, re.I)
        if gl_m:
            gl = gl_m.group(0)
            oe_m = re.search(r'sterreich[:\s]*', gl, re.I)
            if oe_m:
                od = gl[oe_m.end():oe_m.end()+6].lstrip(" |:")
                if _CHECK_PAT.match(od) and not _LEER_PAT.match(od):
                    d["Geburtsland"] = "Österreich"
                else:
                    am = re.search(r'(?:AuR?erhalb|Außerhalb|Auerhalb|Au[&f]erhalb)[^:]*:\s*([^\n]{0,6})', gl, re.I)
                    d["Geburtsland"] = ("Österreich" if (am and LEER.match(am.group(1).strip()))
                                         else "Außerhalb Österreichs")
            else:
                d["Geburtsland"] = "Außerhalb Österreichs"
        else:
            d["Geburtsland"] = "Außerhalb Österreichs"


    # ═══════════════════════════════════════════════════════════════════
    # CHECKBOX-FELDER: Pixel-Dichte primär, Text-Anker als Fallback
    # ═══════════════════════════════════════════════════════════════════

    alle_texte = [_o(doc[i], psm=4) for i in range(len(doc))]
    beh_idx = next((i for i,t in enumerate(alle_texte) if "Behinderung" in t), 1)
    s4i = next((i for i,t in enumerate(alle_texte) if "Minderheit" in t or "Staatszug" in t), 2)
    s4  = alle_texte[s4i]

    # ── Behinderung (Ja/Nein/Keine Angabe) ─────────────────────────────
    beh_pixel = drei_optionen_pixel(doc[beh_idx], ["Behinderung"])
    if beh_pixel is not None:
        d["Behinderung"] = beh_pixel
    else:
        beh_zeile = _zeile_ocr(doc[beh_idx], "Behinderung")
        d["Behinderung"] = _drei_box(beh_zeile) if beh_zeile else "Keine Angabe"

    # ── Minderheit (Ja/Nein/Keine Angabe) ──────────────────────────────
    min_pixel = drei_optionen_pixel(doc[s4i], ["Minderheiten", "marginali"])
    if min_pixel is not None:
        d["Minderheit"] = min_pixel
    else:
        min_zeile = _zeile_ocr(doc[s4i], "Minderheit")
        d["Minderheit"] = _drei_box(min_zeile) if min_zeile else "Keine Angabe"

    # ── Obdachlos (Ja/Nein) ────────────────────────────────────────────
    obd_pixel = zwei_optionen_pixel(doc[s4i], ["Obdachlos"])
    if obd_pixel is not None:
        d["Obdachlos"] = obd_pixel
    else:
        obd_m = re.search(r'Obdachlos.{0,200}', s4, re.DOTALL | re.I)
        obd_zeile = _zeile_ocr(doc[s4i], "Obdachlos")
        obd_txt = (obd_m.group(0) if obd_m else "") + " " + obd_zeile
        d["Obdachlos"] = "Ja" if _drei_box(obd_txt) == "Ja" else "Nein"

    # ── Elternteile im Ausland (Ja/Nein) ───────────────────────────────
    elt_pixel = elternteile_pixel(doc[s4i])
    if elt_pixel is not None:
        d["Elternteile"] = elt_pixel
    else:
        elt_m = re.search(r'(?:Elternteile|Ausland\s+geboren).{0,100}',
                          s4, re.DOTALL | re.I)
        if elt_m:
            d["Elternteile"] = "Ja" if _drei_box(elt_m.group(0)) == "Ja" else "Nein"
        else:
            d["Elternteile"] = "Nein"

    # ── Staatszugehörigkeit (2×2 Grid) ─────────────────────────────────
    # Pixel-Erkennung: alle 4 Optionen separat
    staat_dichten = checkbox_optionen_pixel(
        doc[s4i], ["Staatszugeh"],
        optionen={"sterreich": "Österreich", "schweiz": "EU (ohne Österreich) / EWR / Schweiz"},
        zeilen_toleranz=40)
    staat_dichten2 = checkbox_optionen_pixel(
        doc[s4i], ["Staatenlos"],
        optionen={"staatenlos": "Staatenlos", "drittstaat": "Drittstaat"},
        zeilen_toleranz=40)
    staat_dichten.update(staat_dichten2)

    # Schweiz separat scannen (☒ steht hinter "Schweiz" auf eigener Zeile)
    schweiz_d = checkbox_optionen_pixel(
        doc[s4i], ["Schweiz"],
        optionen={"schweiz": "EU (ohne Österreich) / EWR / Schweiz"},
        zeilen_toleranz=20)
    for k, v in schweiz_d.items():
        staat_dichten[k] = max(staat_dichten.get(k, 0), v)

    staat_gefunden = False
    if staat_dichten:
        gewinner = max(staat_dichten, key=staat_dichten.get)
        if staat_dichten[gewinner] >= 0.16:
            d["Staatszugehörigkeit"] = gewinner
            staat_gefunden = True

    if not staat_gefunden:
        # Text-Fallback (Anker-Methode mit Schweiz/Drittstaat-Zeilen)
        staat_zeile1  = _zeile_ocr(doc[s4i], "Staatszug",  toleranz=30)
        schweiz_zeile = _zeile_ocr(doc[s4i], "Schweiz",    toleranz=15)
        drittst_zeile = _zeile_ocr(doc[s4i], "Drittstaat", toleranz=15)
        staat_m = re.search(r'Staatszugeh.{0,500}', s4, re.DOTALL | re.I)
        staat_txt = " ".join(filter(None, [
            staat_zeile1, schweiz_zeile, drittst_zeile,
            staat_m.group(0) if staat_m else ""]))
        CHK = r"(?:" + _CHECK_PAT.pattern + r")"
        def _sbox(pat):
            m = re.search(pat, staat_txt, re.I)
            if not m: return None
            return staat_txt[m.end():m.end()+6].lstrip(" |:\n\t")
        boxes = {
            "Österreich":                            _sbox(r'sterreich[:\s]*'),
            "EU (ohne Österreich) / EWR / Schweiz":  _sbox(r'Schweiz[\s]*'),
            "Staatenlos":                            _sbox(r'Staatenlos[:\s]*'),
            "Drittstaat":                            _sbox(r'Drittstaat[:\s]*'),
        }
        d["Staatszugehörigkeit"] = "Österreich"
        for wert, box in boxes.items():
            if box and re.search(CHK, box, re.I) and not _LEER_PAT.match(box):
                d["Staatszugehörigkeit"] = wert
                staat_gefunden = True
                break
        if not staat_gefunden:
            hl = {w: (b is None or bool(_LEER_PAT.match(b))) for w, b in boxes.items()}
            ohne = [w for w, l in hl.items() if not l]
            if len(ohne) == 1:
                d["Staatszugehörigkeit"] = ohne[0]
            else:
                d["_staat_unsicher"] = True

    # Eintrittsdatum: scan ALLE Seiten (Formularversion-unabhängig)
    datum_pat_all = r'\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4}'
    d["Eintrittsdatum"] = ""
    for seiten_text in alle_texte:
        m = re.search(r'Datum des Eintritts[^0-9]*(' + datum_pat_all + r')', seiten_text)
        if m:
            d["Eintrittsdatum"] = _normalisiere_datum(m.group(1))
            break
    if not d["Eintrittsdatum"]:
        # Fallback: Unterschriftseite (letztes Datum vor Unterschrift)
        for seiten_text in reversed(alle_texte):
            daten = re.findall(datum_pat_all, seiten_text)
            if daten:
                d["Eintrittsdatum"] = _normalisiere_datum(daten[0])
                break

    doc.close()
    return d


# ═════════════════════════════════════════════════════════════════════════════
# DATENVALIDIERUNG
# ═════════════════════════════════════════════════════════════════════════════

def validiere(daten: dict) -> tuple[bool, list[str]]:
    """
    Prüft die extrahierten Daten auf Vollständigkeit und Widersprüche.
    Gibt (alles_ok, fehlerliste) zurück.
    """
    fehler = []

    # SV-Nummer Widerspruch: Zahl gefunden UND "Keine Angabe" angekreuzt
    if daten.get("SV_Wert_Gefunden") and daten.get("SV_Keine_Angabe"):
        fehler.append(
            "SV-Nummer: Zahl eingetragen aber gleichzeitig 'Keine Angabe' angekreuzt"
        )

    # Adresse und Kontakt
    hat_adresse = any(daten.get(f) for f in ["Straße", "PLZ", "Ort"])
    hat_kontakt = bool(daten.get("Telefon") or daten.get("Email"))

    if not hat_adresse and not hat_kontakt:
        fehler.append("Weder Adresse noch Telefon/Email angegeben")
    elif not hat_adresse:
        fehler.append("Adresse fehlt (Straße/PLZ/Ort leer)")
    elif not hat_kontakt:
        fehler.append("Telefonnummer und E-Mail fehlen")

    # Pflichtfelder
    for feld, name in [
        ("Vorname",       "Vorname"),
        ("Nachname",      "Nachname"),
        ("Geburtsdatum",  "Geburtsdatum"),
        ("Eintrittsdatum","Eintrittsdatum"),
    ]:
        if not daten.get(feld):
            fehler.append(f"{name} konnte nicht extrahiert werden")

    # Staatszugehörigkeit unsicher (Tesseract)
    if daten.get("_staat_unsicher"):
        fehler.append("Staatszugehörigkeit unsicher – bitte manuell prüfen")
    # Claude-OCR Unsicherheiten
    for u in daten.get("_unsicherheiten", []):
        fehler.append(f"OCR unsicher: {u}")
    for p in daten.get("_plausibilitaet", []):
        fehler.append(f"OCR-Problem: {p}")

    return (len(fehler) == 0), fehler


# ═════════════════════════════════════════════════════════════════════════════
# PDF-SUCHE & ORDNER-ERSTELLUNG
# ═════════════════════════════════════════════════════════════════════════════

def finde_pdfs(ordner: Path, vorname: str, nachname: str) -> tuple[Path|None, Path|None]:
    """
    Sucht Sdb- und Zg-PDFs (case-insensitiv) anhand von Vor-/Nachname.
    Gibt (sdb_pfad, zg_pfad) zurück.
    """
    vn = vorname.lower().strip()
    nn = nachname.lower().strip()
    sdb = None
    zg  = None

    for p in ordner.iterdir():
        if p.suffix.lower() != ".pdf":
            continue
        name = p.name.lower()
        # Datei muss Vor- oder Nachname enthalten
        hat_person = (vn in name or nn in name or
                      any(t in name for t in nn.split()) or
                      any(t in name for t in vn.split()))
        if not hat_person:
            continue
        if "sdb" in name or "stamm" in name:
            sdb = p
        elif "zg" in name or "zielgrupp" in name:
            zg = p

    return sdb, zg


def erstelle_ordner_und_verschiebe(
        skript_dir: Path,
        vorname: str,
        nachname: str,
        sdb: Path | None,
        zg:  Path | None
) -> tuple[Path | None, Path | None]:
    """
    Erstellt Ordner 'Vorname Nachname' und verschiebt die PDFs hinein.
    Gibt die neuen Pfade zurück.
    """
    ordner_name = f"{vorname} {nachname}"
    ziel_ordner = skript_dir / ordner_name
    ziel_ordner.mkdir(exist_ok=True)
    log.info("    Ordner: %s", ziel_ordner)

    neuer_sdb = neues_zg = None

    if sdb and sdb.exists():
        neuer_sdb = ziel_ordner / sdb.name
        shutil.move(str(sdb), str(neuer_sdb))
        log.info("    Verschoben: %s", sdb.name)

    if zg and zg.exists():
        neues_zg = ziel_ordner / zg.name
        shutil.move(str(zg), str(neues_zg))
        log.info("    Verschoben: %s", zg.name)

    return neuer_sdb, neues_zg


# ═════════════════════════════════════════════════════════════════════════════
# PLAYWRIGHT-HILFSFUNKTIONEN
# ═════════════════════════════════════════════════════════════════════════════

def ss(page, name: str):
    """Screenshot speichern."""
    SCREENSHOTS.mkdir(exist_ok=True)
    page.screenshot(path=str(SCREENSHOTS / f"{name}.png"))

def fill_s(loc, wert, label, t=5000):
    """Sicher fill – kein Absturz wenn nicht gefunden."""
    if not wert:
        return
    try:
        loc.wait_for(state="visible", timeout=t)
        loc.fill(str(wert))
        log.info("    ✅  %-32s = %s", label, wert)
    except PWTimeout:
        log.warning("    ⚠️   Nicht gefunden: %s", label)


# ═════════════════════════════════════════════════════════════════════════════
# IDEA: LOGIN
# ═════════════════════════════════════════════════════════════════════════════

def login(page, cfg):
    page.goto("https://userapp.idea-esfplus.gv.at/login", wait_until="networkidle")
    page.get_by_role("textbox", name="Benutzername").fill(cfg["idea"]["benutzername"])
    page.get_by_role("textbox", name="Kennwort").fill(cfg["idea"]["passwort"])
    page.get_by_role("button",  name="Anmelden").click()
    page.wait_for_load_state("networkidle", timeout=30000)
    if "login" in page.url.lower():
        raise RuntimeError("Login fehlgeschlagen – Zugangsdaten in config.ini prüfen.")
    log.info("✅ Eingeloggt.")


# ═════════════════════════════════════════════════════════════════════════════
# IDEA: TEILNEHMER ANLEGEN + ID LESEN
# ═════════════════════════════════════════════════════════════════════════════

def lege_teilnehmer_an(page, daten: dict) -> str:
    """
    Legt neuen Teilnehmer via Modal an.
    Liest die Nummer aus dem 'Teilnehmer:innen bearbeiten' Dialog
    (Feld 'Teilnehmer:innen Nummer: XXXX/ YYYY' → gibt YYYY zurück).
    """
    log.info("    Öffne Modal 'Neuer Eintrag'...")
    page.get_by_role("link", name=" Neuer Eintrag").click()
    modal = page.locator("#NeuParticipant")
    modal.wait_for(state="visible", timeout=15000)

    # Namen bereinigen (Unterstriche/OCR-Artefakte entfernen)
    vorname  = re.sub(r'[\s_]+$', '', daten.get("Vorname",  "")).strip()
    nachname = re.sub(r'[\s_]+$', '', daten.get("Nachname", "")).strip()

    try:
        page.locator('#NeuParticipant input[name="TN_Vname"]').fill(vorname)
        log.info("    ✅  %-32s = %s", "Vorname (Modal)", vorname)
    except Exception as e:
        log.warning("    ⚠️   Vorname (Modal) nicht gesetzt: %s", e)

    try:
        page.locator('#NeuParticipant input[name="TN_Nname"]').fill(nachname)
        log.info("    ✅  %-32s = %s", "Nachname (Modal)", nachname)
    except Exception as e:
        log.warning("    ⚠️   Nachname (Modal) nicht gesetzt: %s", e)

    try:
        modal.locator('input[type="radio"]').first.check()
        log.info("    ✅  Radio-Button gesetzt")
    except Exception:
        log.warning("    ⚠️   Radio-Button nicht gefunden")

    ss(page, "02_modal")
    modal.get_by_role("button", name="Speichern").click()
    modal.wait_for(state="detached", timeout=20000)
    page.wait_for_load_state("networkidle", timeout=20000)
    time.sleep(2)
    ss(page, "03_nach_modal")

    # ── Nummer aus "Teilnehmer:innen Nummer: XXXX/ YYYY" lesen ───────────────
    # Das Feld erscheint im 'Allg. Informationen' Tab nach Modal-Speichern
    log.info("    Lese TN-Nummer aus Bearbeiten-Dialog...")

    # Warten bis der Bearbeiten-Dialog mit der Nummer sichtbar ist
    try:
        page.wait_for_selector(
            "text=Teilnehmer:innen Nummer", timeout=20000
        )
    except Exception:
        ss(page, "fehler_keine_nummer")
        raise RuntimeError(
            "Bearbeiten-Dialog nicht geöffnet nach Modal-Speichern.\n"
            "Mögliche Ursachen:\n"
            "  1. IDEA hat den Namen wegen ungültigem Zeichen abgelehnt\n"
            "  2. Session abgelaufen\n"
            f"  Screenshot: screenshots/fehler_keine_nummer.png"
        )

    # Nummer aus readonly-Input lesen (Format: "2661/ 69854")
    nummer = ""
    try:
        # Versuche alle readonly inputs zu durchsuchen
        for loc in page.locator("input[readonly]").all():
            val = loc.input_value()
            m = re.search(r'(\d+)\s*/\s*(\d+)', val)
            if m:
                nummer = m.group(2).strip()
                log.info("    ✅  TN-Nummer gelesen: %s (aus '%s')", nummer, val.strip())
                break
    except Exception as e:
        log.warning("    ⚠️   readonly-Input Scan fehlgeschlagen: %s", e)

    if not nummer:
        # Fallback 2: disabled inputs (IDEA nutzt disabled statt readonly)
        try:
            for loc in page.locator("input[disabled]").all():
                val = loc.input_value()
                m = re.search(r'(\d+)\s*/\s*(\d+)', val)
                if m:
                    nummer = m.group(2).strip()
                    log.info("    ✅  TN-Nummer (disabled-Input): %s", nummer)
                    break
        except Exception as e:
            log.warning("    ⚠️   disabled-Input Scan: %s", e)

    if not nummer:
        # Fallback 3: Seitentext durchsuchen
        try:
            seiten_text = page.locator("body").inner_text()
            m = re.search(r'Teilnehmer:innen Nummer[:\s]+(\d+)\s*/\s*(\d+)', seiten_text)
            if m:
                nummer = m.group(2).strip()
                log.info("    ✅  TN-Nummer (Seitentext): %s", nummer)
        except Exception as e:
            log.warning("    ⚠️   Fallback Textsuche: %s", e)

    if not nummer:
        ss(page, "fehler_keine_nummer")
        raw = input(
            "    📋  Nummer nicht erkannt. Bitte aus Browser kopieren\n"
            "    Format z.B. '2661/ 69854' oder nur '69854': "
        ).strip()
        # Nur die Zahl nach dem "/" nehmen (falls voller String eingegeben)
        if "/" in raw:
            nummer = re.search(r'(\d+)\s*$', raw.split("/")[-1]).group(1).strip()
        else:
            nummer = re.sub(r'\D', '', raw)

    return nummer


def lese_idea_id(page, nummer: str) -> tuple[str, str]:
    """
    Liest die vollständige IDEA-ID aus #tabs_enter nach vollständigen Metadaten.
    Recording: Eintrittsdaten-Tab klicken → #tabs_enter enthält "Name (2619/ 69606)"
    """
    log.info("    Lese IDEA-ID (Eintrittsdaten-Tab erscheint nach erstem Speichern)...")
    # Eintritts-Daten Tab erscheint erst nach erstem Speichern von Allg. Informationen
    # Warte bis "Eintritts-Daten" Tab erscheint (erst nach 1. Speichern sichtbar)
    try:
        page.wait_for_selector(
            'role=tab[name="Eintritts-Daten"]', timeout=20000
        )
        log.info("    🗂️   Eintritts-Daten Tab erschienen")
    except Exception:
        log.warning("    ⚠️   Eintritts-Daten Tab – warte 5s")
        time.sleep(5)
    klicke_tab(page, "Eintritts-Daten")

    try:
        tab_el  = page.locator(f"#tabs_enter{nummer}")
        tab_el.wait_for(state="visible", timeout=10000)
        tab_txt = tab_el.inner_text()
        log.info("    Tab-Text: %s", tab_txt[:80])
        m = re.search(r'\((\d+)\s*/\s*(\d+)\)', tab_txt)
        if m:
            t1, t2 = m.group(1).strip(), m.group(2).strip()
            log.info("    ✅  IDEA-ID: %s / %s", t1, t2)
            return t1, t2
    except Exception as e:
        log.warning("    ⚠️   ID-Text nicht lesbar: %s", e)

    # Fallback: nummer = t2, t1 unbekannt
    raw = input(f"    📋  ID nicht automatisch erkannt. Bitte eingeben (z.B. 2619/69606): ").strip()
    m   = re.search(r'(\d+)\s*/\s*(\d+)', raw)
    return (m.group(1).strip(), m.group(2).strip()) if m else ("", nummer)



def befuelle_interne_kennung(page, nummer: str, interne_kennung: str):
    """
    Trägt die Interne Kennung in #TN_Intern{nummer} ein.
    Wird NACH ID-Lesung aufgerufen (zurück zum Metadaten-Tab).
    """
    if not interne_kennung:
        return
    n = nummer
    log.info("    Trage Interne Kennung nach: %s", interne_kennung)
    try:
        # Metadaten-Tab ist aktiv → direktes .fill() funktioniert
        klicke_tab(page, "Metadaten")
        loc_intern = page.locator(f"#TN_Intern{n}")
        loc_intern.wait_for(state="attached", timeout=8000)
        loc_intern.fill(interne_kennung, force=True)
        log.info("    ✅  %-32s = %s", "Interne Kennung", interne_kennung)
        speichern_und_warten(page, "Interne Kennung")
    except Exception as e:
        log.warning("    ⚠️   Interne Kennung nicht gesetzt: %s", e)




DSGVO_TEXT = (
    "bis 31/07/24 Eva-Maria Himmelbauer/ "
    "ab 01/08/24 Kanzlei Jank Weiler Operenyi I Deloitte Legal"
)


def konvertiere_datum(datum_str: str) -> str:
    """Konvertiert DD.MM.YYYY → YYYY-MM-DD (IDEA-Format)."""
    if not datum_str:
        return ""
    m = re.match(r'(\d{2})[./](\d{2})[./](\d{4})', datum_str.strip())
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return datum_str  # schon im richtigen Format


def _normalisiere_datum(datum_str: str) -> str:
    """
    Normalisiert verschiedene OCR-Datumsformate auf DD.MM.YYYY.
    Akzeptiert: DD.MM.YYYY, DD-MM-YYYY, DD/MM/YYYY, D.M.YY usw.
    """
    if not datum_str:
        return ""
    m = re.match(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})', datum_str.strip())
    if not m:
        return datum_str
    tag, monat, jahr = m.group(1), m.group(2), m.group(3)
    if len(jahr) == 2:
        jahr = "20" + jahr  # 25 → 2025
    return f"{int(tag):02d}.{int(monat):02d}.{jahr}"


def berechne_interne_kennung(t1: str, cfg) -> str:
    """
    Berechnet die Interne Kennung aus Präfix (config.ini) + erstem ID-Teil.
    Beispiel: Präfix="OÖ", t1="2624" → "OÖ 2624"
    """
    prefix = cfg["idea"].get("interne_kennung_prefix", "OÖ").strip()
    return f"{prefix} {t1}" if t1 else ""


def klicke_tab(page, tab_name: str, max_versuche: int = 3):
    """
    Klickt einen Tab und wartet bis er aktiv ist.
    Wiederholt bis zu max_versuche mal falls der Klick nicht registriert wird.
    """
    for versuch in range(1, max_versuche + 1):
        try:
            tab = page.get_by_role("tab", name=tab_name)
            tab.wait_for(state="visible", timeout=10000)
            tab.click()
            # Warte bis Tab-Inhalt geladen ist
            page.wait_for_load_state("domcontentloaded", timeout=10000)
            time.sleep(0.8)
            # Prüfe ob Tab aktiv ist (aria-selected="true")
            try:
                aktiv = tab.get_attribute("aria-selected")
                if aktiv == "true":
                    log.info("    🗂️   Tab '%s' aktiv", tab_name)
                    return True
            except Exception:
                pass
            # Tab sieht aktiv aus auch ohne aria-selected
            if versuch >= max_versuche:
                log.info("    🗂️   Tab '%s' (Versuch %d)", tab_name, versuch)
                return True
        except Exception as e:
            if versuch < max_versuche:
                log.warning("    ⚠️   Tab '%s' Versuch %d: %s – retry...", tab_name, versuch, e)
                time.sleep(1.5)
            else:
                log.warning("    ⚠️   Tab '%s' nicht erreichbar: %s", tab_name, e)
                return False
    return False


def speichern_und_warten(page, label: str = ""):
    """Speichert, wartet auf Stabilität und prüft auf IDEA-Fehlerbanner."""
    try:
        page.get_by_role("button", name="Speichern").click()
        page.wait_for_load_state("networkidle", timeout=15000)
        try:
            page.wait_for_selector(".spinner, .loading", state="hidden", timeout=5000)
        except Exception:
            pass
        time.sleep(0.5)
        # Auf rotes Fehler-Banner prüfen
        try:
            fehler_banner = page.locator("text=Bitte beheben Sie die angeführten Fehler").first
            if fehler_banner.is_visible(timeout=1000):
                log.warning("    ⚠️   IDEA Validierungsfehler nach Speichern! (%s)", label)
                ss(page, f"fehler_speichern_{label.replace(' ','_')}")
        except Exception:
            pass
        if label:
            log.info("    💾 Gespeichert: %s", label)
    except Exception as e:
        log.warning("    ⚠️   Speichern fehlgeschlagen (%s): %s", label, e)


def speichern_warten(page, label=""):
    page.get_by_role("button", name="Speichern").click()
    page.wait_for_load_state("networkidle", timeout=15000)
    time.sleep(1)
    if label:
        log.info("    💾 Gespeichert: %s", label)


def check_s(page, selektor: str, label: str):
    try:
        loc = page.locator(selektor)
        loc.wait_for(state="attached", timeout=4000)
        try:
            loc.check()
        except Exception:
            loc.check(force=True)   # Fallback für versteckte Elemente
        log.info("    ✅  %-32s ✓", label)
    except PWTimeout:
        # Letzter Versuch: direkt mit force
        try:
            page.locator(selektor).check(force=True)
            log.info("    ✅  %-32s ✓ (force)", label)
        except Exception as e:
            log.warning("    ⚠️   Checkbox nicht gefunden: %s  [%s]", label, selektor)


def _panel_reload(page, nummer: str):
    """
    Panel-Reload nach Speichern in 'Allg. Informationen'.
    In der neuen IDEA-Oberfläche gibt es kein #metaEdit mehr –
    stattdessen kurz warten bis die Seite aktualisiert ist.
    """
    try:
        page.wait_for_load_state("networkidle", timeout=5000)
        time.sleep(0.5)
        log.info("    🔄 Panel-Reload (Seite geladen)")
    except Exception:
        time.sleep(1)


def befuelle_hauptformular(page, daten: dict, nummer: str):
    """
    Füllt den Tab 'Allg. Informationen' im Teilnehmer:innen-bearbeiten-Dialog.
    Struktur (aus Screenshots bestätigt):
      1 - Persönliche Daten       → Geschlecht, SV-Nummer
      2 - Wohnort & Kontakt       → Wohnadresse Ja/Nein, Adresse, Telefon, Email
      3 - Rechtliche Rahmenbed.   → 3x Ja-Radio
      4 - Daten-Vollständigkeit   → Ja-Radio (meist schon gesetzt)
      5 - Sonstige Daten          → Datenschutzverantwortliche/r
      → Speichern (einmal am Ende)
    """
    n = nummer

    # ── Tab öffnen ─────────────────────────────────────────────────────────────
    # Metadaten-Tab ist nach dem Modal bereits aktiv – kein Klick nötig
    # Kurz warten bis Formular vollständig geladen
    log.info("    Fülle Metadaten-Tab aus...")
    time.sleep(1.0)

    # ═══════════════════════════════════════════════════════
    # 1 – PERSÖNLICHE DATEN
    # ═══════════════════════════════════════════════════════

    # Geschlecht (22=männlich, 23=weiblich, 24=nicht binär)
    GESCHLECHT_CODES = {"männlich": "22", "weiblich": "23", "nicht binäre person": "24"}
    g_code = GESCHLECHT_CODES.get(daten.get("Geschlecht", "").lower(), "23")
    # Geschlecht: select-Element ist oft hidden → force=True oder JavaScript
    try:
        page.locator(f"#TN_Geschlecht{n}").select_option(g_code, force=True)
        log.info("    ✅  %-32s = %s (Code %s)", "Geschlecht", daten.get("Geschlecht", ""), g_code)
    except Exception:
        try:
            # Fallback: direkt über JavaScript setzen
            page.evaluate(
                f'''(function() {{
                    var s = document.getElementById("TN_Geschlecht{n}");
                    if (s) {{ s.value = "{g_code}"; s.dispatchEvent(new Event("change")); }}
                }})()'''
            )
            log.info("    ✅  Geschlecht (JS) = Code %s", g_code)
        except Exception as e:
            log.warning("    ⚠️   Geschlecht nicht gesetzt: %s", e)

    # SV-Nummer
    sv = daten.get("SV-Nummer", "")
    if sv and sv != "Keine Angabe":
        try:
            loc_sv = page.locator(f"#TN_SVnr{n}")
            loc_sv.wait_for(state="visible", timeout=4000)
            loc_sv.fill(sv)
            log.info("    ✅  %-32s = %s", "SV-Nummer", sv)
        except Exception as e:
            log.warning("    ⚠️   SV-Nummer: %s", e)

    # ═══════════════════════════════════════════════════════
    # 2 – WOHNORT & KONTAKT
    # ═══════════════════════════════════════════════════════

    # "Teilnehmer:in kann Wohnadresse angeben" (TN_Obdachlose)
    # 60 = Ja (hat Adresse)  |  61 = Nein (kein fester Wohnsitz → Telefon reicht)
    hat_adresse = bool(daten.get("Straße", "").strip())
    obd_opt = "60" if hat_adresse else "61"
    check_s(page, f"#TN_Obdachlose{n}{obd_opt}",
            f"Wohnadresse angeben ({'Ja' if hat_adresse else 'Nein – nur Telefon'})")
    time.sleep(1.5)  # IDEA braucht Zeit um Adress-/Telefonfelder ein/auszublenden

    # Adresse (nur wenn vorhanden)
    if hat_adresse:
        for sel, wert, lbl in [
            (f"#TN_Strasse{n}",    daten.get("Straße", ""),  "Straße"),
            (f"#TN_PLZ{n}",        daten.get("PLZ", ""),     "PLZ"),
        ]:
            if wert:
                try:
                    loc = page.locator(sel)
                    loc.wait_for(state="visible", timeout=4000)
                    loc.fill(wert)
                    log.info("    ✅  %-32s = %s", lbl, wert)
                except Exception as e:
                    log.warning("    ⚠️   %s: %s", lbl, e)

        ort = daten.get("Ort", "")
        if ort:
            try:
                page.locator(f"#TN_Ort{n}").select_option(label=ort)
                log.info("    ✅  %-32s = %s", "Ort", ort)
            except Exception as e:
                log.warning("    ⚠️   Ort '%s' nicht im Dropdown: %s", ort, e)

    # Telefon – force=True damit auch nach Obdachlos-Toggle funktioniert
    tel = daten.get("Telefon", "")
    if tel:
        try:
            loc_t = page.locator(f"#TN_Tel{n}")
            loc_t.wait_for(state="attached", timeout=8000)
            loc_t.fill(tel, force=True)
            log.info("    ✅  %-32s = %s", "Telefonnummer", tel)
        except Exception as e:
            try:
                # JS-Fallback: direkt ins DOM
                page.evaluate(f"""
                    var el = document.getElementById('TN_Tel{n}');
                    if (!el) el = document.querySelector('input[name="TN_Tel"]');
                    if (el) {{
                        el.value = '{tel}';
                        el.dispatchEvent(new Event('input',  {{bubbles: true}}));
                        el.dispatchEvent(new Event('change', {{bubbles: true}}));
                    }}
                """)
                log.info("    ✅  Telefon (JS) = %s", tel)
            except Exception as e2:
                log.warning("    ⚠️   Telefon: %s", e2)

    # E-Mail
    email = daten.get("Email", "")
    if email:
        try:
            loc_m = page.locator(f"#TN_Email{n}")
            loc_m.wait_for(state="visible", timeout=4000)
            loc_m.fill(email)
            log.info("    ✅  %-32s = %s", "E-Mail", email)
        except Exception as e:
            log.warning("    ⚠️   E-Mail: %s", e)

    # ═══════════════════════════════════════════════════════
    # 3 – RECHTLICHE RAHMENBEDINGUNGEN  (alle Ja = Option 60)
    # ═══════════════════════════════════════════════════════
    for feld, opt, lbl in [
        ("TN_Datverar_ok", "60", "Daten weiterverwendet (Ja)"),
        ("TN_Dat_ok",      "60", "Richtigkeit bestätigt (Ja)"),
        ("TN_Publ_ok",     "60", "JTF/ESF+ Kenntnis (Ja)"),
    ]:
        check_s(page, f"#{feld}{n}{opt}", lbl)

    # ═══════════════════════════════════════════════════════
    # 4 – DATEN-VOLLSTÄNDIGKEIT  (Ja = Option 60)
    # ═══════════════════════════════════════════════════════
    check_s(page, f"#TN_Incomp{n}60", "Daten vollständig (Ja)")

    # ═══════════════════════════════════════════════════════
    # 5 – SONSTIGE DATEN
    # ═══════════════════════════════════════════════════════

    # Datenschutzverantwortliche/r  (click aktiviert, dann fill)
    try:
        loc_dsgvo = page.locator(f"#TN_DSGVO{n}")
        loc_dsgvo.wait_for(state="visible", timeout=6000)
        loc_dsgvo.click()
        loc_dsgvo.fill(DSGVO_TEXT)
        log.info("    ✅  %-32s = %s…", "Datenschutzverantwortliche/r", DSGVO_TEXT[:40])
    except Exception as e:
        try:
            page.locator('input[name="TN_DSGVO"], textarea[name="TN_DSGVO"]').fill(DSGVO_TEXT)
            log.info("    ✅  DSGVO (Fallback)")
        except Exception as e2:
            log.warning("    ⚠️   DSGVO: %s", e2)

    # ── Einmal Speichern am Ende ────────────────────────────────────────────────
    speichern_und_warten(page, "Metadaten")
    ss(page, f"05_metadaten_fertig_{n}")
    log.info("    ✅  Metadaten vollständig gespeichert.")


def befuelle_eintrittsdaten(page, daten: dict, nummer: str):
    """
    Füllt den Tab Eintritts-Daten – gleiche Logik wie Allg. Informationen:
    alle Felder befüllen, einmal Speichern am Ende.
    """
    n = nummer
    log.info("    Fülle Eintritts-Daten aus...")

    # ── Förderrelevanter Eintritt = Ja ────────────────────────────────────
    check_s(page, f"#FR_EI{n}61", "Förderrelevanter Eintritt (Ja)")

    # ── Eintrittsdatum ────────────────────────────────────────────────────
    eintritt_idea = konvertiere_datum(daten.get("Eintrittsdatum", ""))
    if eintritt_idea:
        try:
            page.locator(f"#FR_EI_Dat{n}").fill(eintritt_idea, force=True)
            log.info("    ✅  %-32s = %s", "Eintrittsdatum", eintritt_idea)
        except Exception as e:
            log.warning("    ⚠️   Eintrittsdatum: %s", e)

    # ── Geburtsland ───────────────────────────────────────────────────────
    GEBLAND_CODES = {"österreich": "37", "außerhalb österreichs": "38"}
    gl_code = GEBLAND_CODES.get(daten.get("Geburtsland", "").lower(), "38")
    try:
        page.locator(f"#TN_GebLand{n}").select_option(gl_code, force=True)
        log.info("    ✅  %-32s = %s (Code %s)",
                 "Geburtsland", daten.get("Geburtsland", ""), gl_code)
    except Exception as e:
        log.warning("    ⚠️   Geburtsland: %s", e)

    # ── Geburtsjahr bekannt + Geburtsdatum ────────────────────────────────
    check_s(page, f"#TN_GebJahr_JN{n}61", "Geburtsjahr bekannt (Ja)")
    geb_idea = konvertiere_datum(daten.get("Geburtsdatum", ""))
    if geb_idea:
        try:
            page.locator(f"#TN_GebDat{n}").fill(geb_idea, force=True)
            log.info("    ✅  %-32s = %s", "Geburtsdatum", geb_idea)
        except Exception as e:
            log.warning("    ⚠️   Geburtsdatum: %s", e)

    # ── EECO12: Behinderung ───────────────────────────────────────────────
    BEHINDERUNG_CODES = {"Ja": "62", "Nein": "63", "Keine Angabe": "64"}
    beh_code = BEHINDERUNG_CODES.get(daten.get("Behinderung", "Keine Angabe"), "64")
    check_s(page, f"#EECO12{n}{beh_code}",
            f"Behinderung ({daten.get('Behinderung','Keine Angabe')})")

    # ── Feste EECO-Werte ──────────────────────────────────────────────────
    try:
        page.locator(f"#EECO09H{n}").select_option("48", force=True)
        log.info("    ✅  EECO09H = 48 (fix)")
    except Exception:
        try:
            page.locator('select[name="EECO09H"]').select_option("48", force=True)
        except Exception as e:
            log.warning("    ⚠️   EECO09H: %s", e)

    # EECO20H = immer Option 60 (fix, bestätigt aus Recording)
    check_s(page, f"#EECO20H{n}60", "EECO20H (fix=60)")

    try:
        page.locator(f"#EECO02H{n}").select_option("43", force=True)
        log.info("    ✅  EECO02H = 43 (fix)")
    except Exception:
        try:
            page.locator('select[name="EECO02H"]').select_option("43", force=True)
        except Exception as e:
            log.warning("    ⚠️   EECO02H: %s", e)

    # ── EECO15: Minderheit ────────────────────────────────────────────────
    MINDERHEIT_CODES = {"Ja": "62", "Nein": "63", "Keine Angabe": "64"}
    min_code = MINDERHEIT_CODES.get(daten.get("Minderheit", "Keine Angabe"), "64")
    check_s(page, f"#EECO15{n}{min_code}",
            f"Minderheit ({daten.get('Minderheit','Keine Angabe')})")

    # ── EECO16: Wohnungssituation ─────────────────────────────────────────
    eeco16_code = "61" if daten.get("Obdachlos", "Nein") == "Nein" else "60"
    check_s(page, f"#EECO16{n}{eeco16_code}",
            f"EECO16 Wohnungssituation ({daten.get('Obdachlos','Nein')})")

    # ── EECO14: Beide Elternteile im Ausland geboren ─────────────────────
    # 62 = Ja  |  63 = Nein  (✅ bestätigt aus Recording)
    ELTERNTEILE_CODES = {"Ja": "62", "Nein": "63"}
    elt_code = ELTERNTEILE_CODES.get(daten.get("Elternteile", "Nein"), "63")
    check_s(page, f"#EECO14{n}{elt_code}",
            f"Elternteile im Ausland ({daten.get('Elternteile','Nein')})") 

    # ── EECO13H: Staatszugehörigkeit ──────────────────────────────────────
    STAAT_CODES = {
        "Österreich": "6",
        "EU (ohne Österreich) / EWR / Schweiz": "7",
        "Drittstaat": "8",
        "Staatenlos": "9",
    }
    staat_code = STAAT_CODES.get(daten.get("Staatszugehörigkeit", ""), "7")
    try:
        page.locator(f"#EECO13H{n}").select_option(staat_code, force=True)
        log.info("    ✅  %-32s = %s (Code %s)",
                 "Staatszugehörigkeit", daten.get("Staatszugehörigkeit", ""), staat_code)
    except Exception:
        try:
            page.locator('select[name="EECO13H"]').select_option(staat_code, force=True)
        except Exception as e:
            log.warning("    ⚠️   EECO13H: %s", e)

    # ── Einmal Speichern ──────────────────────────────────────────────────
    speichern_und_warten(page, "Eintrittsdaten")
    ss(page, f"06_eintrittsdaten_fertig_{n}")
    log.info("    ✅  Eintrittsdaten vollständig gespeichert.")


def main():
    log.info("═" * 60)
    log.info("  ESF+ Pipeline gestartet")
    log.info("═" * 60)

    if not CONFIG_FILE.exists():
        sys.exit(f"FEHLER: config.ini nicht gefunden: {CONFIG_FILE}")

    cfg = configparser.ConfigParser()
    cfg.read(CONFIG_FILE, encoding="utf-8")

    excel_pfad = Path(cfg["pfade"]["excel_ausgabe"])
    skript_dir = Path(__file__).parent

    if not excel_pfad.exists():
        sys.exit(f"FEHLER: Excel nicht gefunden: {excel_pfad}")

    # Excel laden (zum Schreiben)
    wb = load_workbook(str(excel_pfad))
    ws = wb.active

    # Offene Zeilen ermitteln
    zeilen = [
        r for r in range(2, ws.max_row + 1)
        if str(ws.cell(row=r, column=COL_STATUS).value or "").strip() == STATUS_BEREIT
    ]

    if not zeilen:
        log.info("Keine offenen Einträge (Spalte E = '%s').", STATUS_BEREIT)
        return

    log.info("%d Einträge zu bearbeiten.", len(zeilen))

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, slow_mo=150)
        context = browser.new_context(viewport={"width": 1440, "height": 900})
        page    = context.new_page()

        # Einmalig einloggen + Verwaltung öffnen
        log.info("── Login ──")
        login(page, cfg)
        page.get_by_role("link", name=" Teilnehmer:innen-Verwaltung").click()
        page.wait_for_load_state("networkidle", timeout=20000)
        ss(page, "01_verwaltung")

        for zeile in zeilen:
            vorname  = str(ws.cell(row=zeile, column=COL_VORNAME).value  or "").strip()
            nachname = str(ws.cell(row=zeile, column=COL_NACHNAME).value or "").strip()

            log.info("")
            log.info("── Zeile %d: %s %s ──", zeile, vorname, nachname)

            # ── 1. PDFs finden ────────────────────────────────────────────────
            sdb, zg = finde_pdfs(skript_dir, vorname, nachname)

            if not sdb:
                log.error("    Kein Stammdatenblatt-PDF gefunden – übersprungen.")
                ws.cell(row=zeile, column=COL_OK,     value="Nein")
                ws.cell(row=zeile, column=COL_FEHLER, value="Kein Stammdatenblatt-PDF gefunden")
                wb.save(str(excel_pfad))
                continue

            if not zg:
                log.warning("    Kein Zielgruppennachweis-PDF gefunden – nur Sdb.")

            # ── 2. Ordner anlegen + PDFs verschieben ──────────────────────────
            neuer_sdb, neues_zg = erstelle_ordner_und_verschiebe(
                skript_dir, vorname, nachname, sdb, zg
            )

            # ── 3. OCR ────────────────────────────────────────────────────────
            try:
                daten = parse_stammdatenblatt(neuer_sdb)
            except Exception as e:
                log.error("    OCR-Fehler: %s", e)
                ws.cell(row=zeile, column=COL_OK,     value="Nein")
                ws.cell(row=zeile, column=COL_FEHLER, value=f"OCR-Fehler: {e}")
                wb.save(str(excel_pfad))
                continue

            # ── 4. Daten validieren ───────────────────────────────────────────
            ok, fehler_liste = validiere(daten)
            if not ok:
                log.warning("    ⚠️  Datenfehler: %s", " | ".join(fehler_liste))
                ws.cell(row=zeile, column=COL_OK,     value="Nein")
                ws.cell(row=zeile, column=COL_FEHLER, value=" | ".join(fehler_liste))
            else:
                ws.cell(row=zeile, column=COL_OK,     value="Ja")
                ws.cell(row=zeile, column=COL_FEHLER, value="")
            wb.save(str(excel_pfad))

            # ── 5. IDEA: Modal → Person anlegen, Nummer holen ────────────────────
            try:
                nummer = lege_teilnehmer_an(page, daten)
            except (KeyboardInterrupt, SystemExit):
                log.info("Abgebrochen.")
                break
            except Exception as e:
                log.error("    IDEA-Fehler beim Anlegen: %s", e)
                ss(page, f"fehler_zeile_{zeile}")
                weiter = input("    ⚠️  Fehler – Zeile überspringen? (j/n): ").strip().lower()
                if weiter != "j":
                    break
                continue

            # ── 6. Metadaten ausfüllen (Obdachlos Nein + alle Felder, ohne Interne Kennung)
            # WICHTIG: Obdachlos "Nein" MUSS gesetzt sein, sonst kein Speichern möglich
            befuelle_hauptformular(page, daten, nummer)

            # ── 7. Eintrittsdaten-Tab → IDEA-ID lesen ────────────────────────
            # ID erscheint erst nach vollständig gespeicherten Metadaten
            try:
                t1, t2 = lese_idea_id(page, nummer)
            except Exception as e:
                log.error("    ID-Lesung fehlgeschlagen: %s", e)
                t1, t2 = "", nummer

            # ── 8. Excel: R + S eintragen ────────────────────────────────────
            ws.cell(row=zeile, column=COL_ID_1, value=t1)
            ws.cell(row=zeile, column=COL_ID_2, value=t2)
            wb.save(str(excel_pfad))

            # ── 9. Interne Kennung berechnen → eintragen ─────────────────────
            interne_kennung = berechne_interne_kennung(t1, cfg)
            log.info("    Interne Kennung: %s", interne_kennung)
            befuelle_interne_kennung(page, nummer, interne_kennung)

            # ── 10. Zu Eintritts-Daten navigieren ────────────────────────────
            klicke_tab(page, "Eintritts-Daten")

            # ── 11. Eintrittsdaten ausfüllen ──────────────────────────────────
            befuelle_eintrittsdaten(page, daten, nummer)

            # ── 11. Abschlusskontrolle ─────────────────────────────────────────
            print(f"\n{'─'*55}")
            print(f"  👤  {daten['Vorname']} {daten['Nachname']}")
            print(f"  📅  Geburt: {daten.get('Geburtsdatum','—')}  Eintritt: {daten.get('Eintrittsdatum','—')}")
            print(f"  📊  Erwerbsstatus: {daten.get('Erwerbsstatus','—')}")
            print(f"  🔑  IDEA-ID: {t1}/{t2}  |  Interne Kennung: {interne_kennung}")
            if not ok:
                print(f"  ⚠️   Datenfehler: {' | '.join(fehler_liste)}")
            print(f"{'─'*55}")
            print("  ✅  Alle Felder automatisch eingetragen.")
            print("  👉  Bitte Datensatz in IDEA kontrollieren.")
            input("  👉  ENTER drücken um fortzufahren...")

            ss(page, f"05_gespeichert_{nachname}")

            # ── 12. Excel: Status + grün einfärben ────────────────────────────
            ws.cell(row=zeile, column=COL_STATUS, value=STATUS_ERLEDIGT)
            wb.save(str(excel_pfad))

            log.info("    💾 Zeile %d abgeschlossen → '%s'", zeile, STATUS_ERLEDIGT)


        # Browser sauber schließen
        try:
            context.close()
            browser.close()
        except Exception:
            pass
    log.info("")
    log.info("═" * 60)
    log.info("  Pipeline abgeschlossen.")
    log.info("═" * 60)


if __name__ == "__main__":
    main()
